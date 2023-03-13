import openpyxl
from openpyxl import load_workbook
import datetime

workbook = openpyxl.load_workbook(
    'C:/Users/JFROJAS/Desktop/Barcode/src/exampledb.xlsx')
sheet = workbook.active

if __name__ == '__main__':
    # crear archivo de resultados con nombre único
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
    resultados_filename = f"resultados_{timestamp}.xlsx"
    resultados_workbook = openpyxl.Workbook()
    resultados_sheet = resultados_workbook.active

    # agregar encabezados al archivo de resultados
    encabezados = ['Nombre', 'Clave', 'Lote', 'Fecha de caducidad']
    for i, encabezado in enumerate(encabezados, start=1):
        resultados_sheet.cell(row=1, column=i, value=encabezado)

    while True:
        barcode = int(input("Ingrese el código de barras a buscar: "))
        if barcode == 500:
            break
        name = []
        clave = []
        lote = []
        fecha_caducidad = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == barcode:
                name.append(row[1])
                clave.append(row[2])
                # asumiendo que el número del lote está en la columna 4
                lote.append(row[3])
                # asumiendo que la fecha de caducidad está en la columna 5
                fecha_caducidad.append(row[4])

        # Imprimir resultados
        if name:
            print("Nombre(s):", ", ".join(name))
            print("Clave(s):", ", ".join(clave))
            # mostrar opciones únicas de lote
            print("Lote(s):", ", ".join(set(lote)))
            lote_seleccionado = input("Seleccione un lote: ")
            # encontrar las filas correspondientes al lote seleccionado
            resultados = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == barcode and row[3] == lote_seleccionado:
                    resultados.append({
                        'Nombre': row[1],
                        'Clave': row[2],
                        'Lote': row[3],
                        'Fecha de caducidad': row[4]
                    })
            # imprimir resultados
            print(f"Resultados para el lote {lote_seleccionado}:")
            for resultado in resultados:
                print("Nombre:", resultado['Nombre'])
                print("Clave:", resultado['Clave'])
                print("Lote:", resultado['Lote'])
                print("Fecha de caducidad:", resultado['Fecha de caducidad'])
            # agregar resultados al archivo de resultados
            for resultado in resultados:
                resultados_sheet.append(
                    [resultado['Nombre'], resultado['Clave'], resultado['Lote'], resultado['Fecha de caducidad']])
        else:
            print("No se encontraron resultados")

        # guardar archivo de resultados con nombre único
        resultados_workbook.save(resultados_filename)
