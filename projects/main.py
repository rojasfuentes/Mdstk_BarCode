import openpyxl
from openpyxl import load_workbook
import datetime
import tkinter as tk
from tkinter import messagebox

# Abrir archivo de excel y cargar hoja activa
workbook = openpyxl.load_workbook(
    'C:/Users/JFROJAS/Desktop/Barcode/src/exampledb.xlsx')
sheet = workbook.active

# Crear archivo de resultados y cargar hoja activa
timestamp = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
resultados_filename = f"resultados_{timestamp}.xlsx"
resultados_workbook = openpyxl.Workbook()
resultados_sheet = resultados_workbook.active

# Agregar encabezados al archivo de resultados
encabezados = ['Barcode', 'Nombre', 'Clave', 'Lote', 'Fecha de caducidad']
for i, encabezado in enumerate(encabezados, start=1):
    resultados_sheet.cell(row=1, column=i, value=encabezado)


def salir():
    root.destroy()


def search_barcode(barcode):
    name = []
    clave = []
    lotes = []
    fecha_caducidad = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == barcode:
            name.append(row[1])
            clave.append(row[2])
            lotes.append(row[3])
            fecha_caducidad.append(row[4])
    if not name:
        messagebox.showerror("Error", "No se encontraron resultados")
    else:
        lotes_seleccionados = list(set(lotes))
        lotes_seleccionados.sort()
        seleccion_lotes(lotes_seleccionados, barcode, name,
                        clave, lotes, fecha_caducidad)


def seleccion_lotes(lotes_seleccionados, barcode, name, clave, lotes, fecha_caducidad):
    tk.Label(root, text="Seleccione un lote").pack()
    for lote in lotes_seleccionados:
        tk.Button(root, text=lote, command=lambda lote_seleccionado=lote: resultados_busqueda(
            lote_seleccionado, barcode, name, clave, lotes, fecha_caducidad)).pack()


def resultados_busqueda(lote, barcode, name, clave, lotes, fecha_caducidad):
    resultados = []
    for row in range(len(lotes)):
        if lotes[row] == lote:
            resultados.append({
                'Barcode': barcode,
                'Nombre': name[row],
                'Clave': clave[row],
                'Lote': lote,
                'Fecha de caducidad': fecha_caducidad[row],
                'Cantidad': 0
            })
    agregar_cantidad(resultados)


def agregar_cantidad(resultados):
    tk.Label(
        root, text="Cantidad de productos:").pack()
    cantidad_var = tk.StringVar()  # Crear objeto StringVar
    # Asignar objeto StringVar a la Entry
    cantidad_entry = tk.Entry(root, textvariable=cantidad_var)
    cantidad_entry.pack()

    def guardar():
        cantidad = int(cantidad_var.get())  # Convertir entrada a int
        resultados[0]['Cantidad'] = cantidad
        guardar_resultados(resultados)
    # Botón para guardar cantidad y llamar a guardar_resultados
    tk.Button(root, text="Guardar", command=guardar).pack()


def guardar_resultados(resultados):
    # agregar resultados al archivo de resultados
    for resultado in resultados:
        resultados_sheet.append(
            [resultado['Barcode'], resultado['Nombre'], resultado['Clave'], resultado['Lote'], resultado['Fecha de caducidad'], resultado['Cantidad']])

    # guardar archivo de resultados
    resultados_workbook.save(resultados_filename)
    messagebox.showinfo(
        "Éxito", f"Se han encontrado {len(resultados)} resultado(s) para el lote seleccionado")
    reiniciar_interfaz()


def reiniciar_interfaz():
    # Borra todos los widgets de la ventana
    for widget in root.winfo_children():
        widget.destroy()

    # Vuelve a crear los widgets necesarios
    label = tk.Label(root, text="Ingrese el código de barras a buscar:")
    label.pack()
    barcode_entry = tk.Entry(root)
    barcode_entry.pack()

    def on_click():
        barcode = int(barcode_entry.get())
        barcode_entry.delete(0, tk.END)
        try:
            search_barcode(barcode)
        except ValueError:
            messagebox.showerror(
                "Error", "El valor ingresado no es un número entero válido")

    tk.Button(root, text="Buscar", command=on_click).pack()
    tk.Button(root, text="Salir", command=salir).pack()

    # Restablece la ventana a su tamaño mínimo
    root.geometry("")


if __name__ == '__main__':
    root = tk.Tk()
    root.focus_force()
    root.title("Buscar código de barras")
    label = tk.Label(root, text="Ingrese el código de barras a buscar:")
    label.pack()
    barcode_entry = tk.Entry(root)
    barcode_entry.pack()
    barcode_entry.focus()

    def on_click():
        barcode = int(barcode_entry.get())
        barcode_entry.delete(0, tk.END)
        try:
            search_barcode(barcode)
        except ValueError:
            messagebox.showerror(
                "Error", "El valor ingresado no es un número entero válido")

    tk.Button(root, text="Buscar", command=on_click).pack()

    tk.Button(root, text="Salir", command=salir).pack()
    root.mainloop()
